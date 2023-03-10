import datetime
import asyncio

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Side, Border
# lookFor = '990110162483870'
# def data_by_trek(lookFor):
#     for i in range(1, sheet.max_row+1):
#         value = sheet.cell(row=i, column=4).value
#         if value == lookFor:
#             availble = sheet[i][10].value
#             price = sheet[i][11].value
#             trek = sheet[i][3].value
#             position = sheet[i][5].value
#             return [trek, availble, position, price]

# print(data_by_trek(lookFor=lookFor))

# name2 = '294'
# name3 = 'GX-0294'


# async def data_by_id(id1, id2, id3, id4):
#     wb_search = load_workbook('data.xlsx')
#     sheet = wb_search.active
#     data = []
#     for i in range(1, sheet.max_row + 1):
#         data_dict = {}
#         value = str(sheet.cell(row=i, column=3).value)
#         if id1 == value or id2 == value or id3 == value or id4 == value:
#             data_dict['no'] = id1
#             data_dict['no_auto'] = sheet[i][0].value
#             try:
#                 data_dict['date'] = sheet[i][10].value.strftime('%d.%m.%Y')
#             except AttributeError:
#                 data_dict['date'] = None
#
#             data_dict['trek'] = sheet[i][3].value
#
#             data_dict['weight'] = sheet[i][7].value
#             data_dict['price'] = sheet[i][8].value
#             data_dict['in_china'] = sheet[i][6].value
#             if data_dict.get('in_china') is None:
#                 data_dict['in_china'] = '0'
#             if data_dict.get('weight') is None:
#                 data_dict['weight'] = '0'
#             if data_dict.get('price') is None:
#                 data_dict['price'] = '0'
#             data_dict['sum'] = int(data_dict.get('price')) * int(data_dict.get('weight')) + int(data_dict.get(
#                 'in_china'))
#
#             data_dict['opl'] = sheet[i][11].value
#             data.append(data_dict)
#
#     for dictionary in data:
#         for key, value in dictionary.items():
#             if value is None:
#                 dictionary[key] = 'Не указано'
#
#     if len(data) > 20:
#         big_data = []
#         str_to_append = ''
#         for i in range(len(data)):
#             str_to_append += (f'{i + 1} товар\n'
#                               f'Номер товара: {data[i]["no"]}\n'
#                               f'Номер авто: {data[i]["no_auto"]}\n'
#                               f'Дата выдачи: {data[i]["date"]}\n'
#                               f'Трек: {data[i]["trek"]}\n'
#                               f'Вес: {data[i]["weight"]}\n'
#                               f'Цена: {data[i]["price"]}\n'
#                               f'Сумма: {data[i]["sum"]}\n'
#                               f'Статус: {data[i]["opl"]}\n\n')
#             if i % 20 == 0 and i != 0:
#                 big_data.append(str_to_append)
#                 str_to_append = ''
#         return big_data
#
#     else:
#         small_data = []
#         str_to_append = ''
#         for i in range(len(data)):
#             str_to_append += (f'{i + 1} товар\n'
#                               f'Номер товара: {data[i]["no"]}\n'
#                               f'Номер авто: {data[i]["no_auto"]}\n'
#                               f'Дата выдачи: {data[i]["date"]}\n'
#                               f'Трек: {data[i]["trek"]}\n'
#                               f'Вес: {data[i]["weight"]}\n'
#                               f'Цена: {data[i]["price"]}\n'
#                               f'Сумма: {data[i]["sum"]}\n'
#                               f'Статус: {data[i]["opl"]}\n\n')
#         small_data.append(str_to_append)
#         return small_data

# print(data_by_id('75', 'GX-0075', '21', '212'))


#         {'Модель': f'GX-{sheet[i][2].value}','ТРЕК': sheet[i][3].value,
#         'Коробка': sheet[i][4].value, 'Позиция': sheet[i][5].value,
#         'В китае': sheet[i][6].value, 'Вес': sheet[i][7].value,
#         'Цена': sheet[i][8].value, 'Сумма': sheet[i][9].value,
#         'Дата выдачи': sheet[i][10].value, 'Оплачено': sheet[i][11].value,
#         'Примичание': sheet[i][12].value, 'Расчет': sheet[i][13].value,
#         'Полка': sheet[i][14].value}


async def data_by_trek(trek, delay):
    await asyncio.sleep(delay)
    wb_search = load_workbook('/home/Saynir/bot/data.xlsx')
    sheet = wb_search.active
    data = []
    for i in range(1, sheet.max_row + 1):
        lst = []
        value = str(sheet.cell(row=i, column=4).value)
        if trek == value:
            lst.append(sheet[i][2].value)
            lst.append(sheet[i][0].value)
            lst.append(sheet[i][1].value)
            # try:
            #     lst.append(sheet[i][1].value.strftime('%d.%m.%Y'))
            # except AttributeError:
            #     lst.append(None)

            lst.append(sheet[i][3].value)

            lst.append(sheet[i][4].value)
            lst.append(sheet[i][5].value)
            lst.append(sheet[i][6].value)
            if lst[6] is None:
                lst[6] = 0
            if lst[4] is None:
                lst[4] = 0
            if lst[5] is None:
                lst[5] = 0

            lst.append(sheet[i][8].value)
            data.append(lst)

    # for dictionary in data:
    #     for key, value in dictionary.items():
    #         if value is None:
    #             dictionary[key] = 'Не указано'
    if len(data) == 0:
        return 1
    wb = load_workbook('/home/Saynir/bot/Ваши_товары.xlsx')
    sheet = wb.active
    clear(wb, sheet)
    a = 0
    for i in range(1, len(data)+2):
        if i == 1:
            for c in range(9):
                side = Side(border_style='thin')
                border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side,
                )
                sheet[i][c].border = border
            continue
        if a == len(data)+1:
            d = data[a]
            sheet[i][0].value = d[0]
            sheet[i][1].value = d[2]
            sheet[i][2].value = str(d[3])
            sheet[i][3].value = d[4]
            sheet[i][4].value = d[5]
            sheet[i][5].value = d[6]
            sheet[i][6].value = d[7]
            sheet[i][7].value = d[1]
            break
        else:
            d = data[a]
            sheet[i][0].value = d[0]
            sheet[i][1].value = d[2]
            sheet[i][2].value = str(d[3])
            sheet[i][3].value = d[4]
            sheet[i][4].value = d[5]
            try:
                if float(d[4]) != 0.0 or float(d[5]) != 0.0:
                    sum = (d[4] * d[5] + d[6])
                    sheet[i][6].value = round(sum, 2)
                else:
                    sheet[i][6].value = 0
            except Exception:
                pass

            sheet[i][5].value = d[6]
            sheet[i][7].value = d[7]
            for c in range(9):
                side = Side(border_style='thin')
                border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side,
                )
                sheet[i][c].border = border
                if d[7].upper() == 'ПОЛУЧЕНО':
                    sheet[i][c].fill = PatternFill(
                        start_color='1BEC11', end_color='1BEC11', fill_type="solid")
                elif d[7].upper() == 'БИШКЕК':
                    sheet[i][c].fill = PatternFill(
                        start_color='FFF700', end_color='FFF700', fill_type="solid")
                elif d[7].upper() == 'ОТПРАВЛЕН ОШ':
                    sheet[i][c].fill = PatternFill(
                        start_color='0047AB', end_color='0047AB', fill_type="solid")
                else:
                    sheet[i][c].fill = PatternFill(
                        start_color='ED1717', end_color='ED1717', fill_type="solid")
            sheet[i][8].value = d[1]
            a += 1
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['H'].width = 18
    wb.save('/home/Saynir/bot/Ваши_товары.xlsx')


def data_to_notifications():
    wb_search = load_workbook('bot/data.xlsx')
    sheet = wb_search.active
    gx_lst = []
    trek_lst = []
    for i in range(1, sheet.max_row + 1):
        value = sheet.cell(row=i, column=11).value
        if str(value)[2:10] == datetime.datetime.now().strftime('%Y-%m-%d'):
            gx_lst.append(sheet[i][2].value)

    gx_lst = list(set(gx_lst))

    for i in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row=i, column=3).value)
        for n in gx_lst:
            if value == str(n):
                trek_lst.append((n, sheet[i][3].value))
    return trek_lst


async def dataXlsx(id1, delay):
    await asyncio.sleep(delay)
    print(id1)
    wb_search = load_workbook('/home/Saynir/bot/data.xlsx')
    sheet = wb_search.active
    data = []
    for i in range(1, sheet.max_row + 1):
        data_dict = []
        value = str(sheet.cell(row=i, column=3).value)
        # print(value)
        # print(value==id1)
        if id1 == value:
            print(value)
            data_dict.append(id1)
            data_dict.append(sheet[i][0].value)
            data_dict.append(sheet[i][1].value)
            # try:
            #     data_dict.append(sheet[i][1].value.strftime('%d.%m.%Y'))
            # except AttributeError:
            #     data_dict.append(None)

            data_dict.append(sheet[i][3].value)

            data_dict.append(sheet[i][4].value)
            data_dict.append(sheet[i][5].value)
            data_dict.append(sheet[i][6].value)
            if data_dict[6] is None:
                data_dict[6] = 0
            if data_dict[4] is None:
                data_dict[4] = 0
            if data_dict[5] is None:
                data_dict[5] = 0

            data_dict.append(sheet[i][8].value)
            data.append(data_dict)
    print(data)
    print(len(data) == 0)
    if len(data) == 0:
        return 1
    wb = load_workbook('/home/Saynir/bot/Ваши_товары.xlsx')
    sheet = wb.active
    clear(wb, sheet)
    a = 0
    # try:
    for i in range(1, len(data)+2):
        if i == 1:
            for c in range(9):
                side = Side(border_style='thin')
                border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side,
                )
                sheet[i][c].border = border
            continue
        if a == len(data)+1:
            d = data[a]
            sheet[i][0].value = d[0]
            sheet[i][1].value = d[2]
            sheet[i][2].value = str(d[3])
            sheet[i][3].value = d[4]
            sheet[i][4].value = d[5]
            sheet[i][5].value = d[6]
            sheet[i][6].value = d[7]
            sheet[i][7].value = d[8]
            sheet[i][7].value = d[1]
            break
        else:
            d = data[a]
            sheet[i][0].value = d[0]
            sheet[i][1].value = d[2]
            sheet[i][2].value = str(d[3])
            sheet[i][3].value = d[4]
            sheet[i][4].value = d[5]
            try:
                if float(d[4]) != 0.0 or float(d[5]) != 0.0:
                    sum = (d[4] * d[5] + d[6])
                    sheet[i][6].value = round(sum, 2)
                else:
                    sheet[i][6].value = 0
            except Exception:
                pass

            sheet[i][5].value = d[6]
            sheet[i][7].value = d[7]
            for c in range(9):
                side = Side(border_style='thin')
                border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side,
                )
                sheet[i][c].border = border
                if not d[7] is None:
                    if d[7].upper() == 'ПОЛУЧЕНО':
                        sheet[i][c].fill = PatternFill(
                            start_color='1BEC11', end_color='1BEC11', fill_type="solid")
                    elif d[7].upper() == 'БИШКЕК':
                        sheet[i][c].fill = PatternFill(
                            start_color='FFF700', end_color='FFF700', fill_type="solid")
                    elif d[7].upper() == 'ОТПРАВЛЕН ОШ':
                        sheet[i][c].fill = PatternFill(
                            start_color='0047AB', end_color='0047AB', fill_type="solid")
                    else:
                        sheet[i][c].fill = PatternFill(
                            start_color='ED1717', end_color='ED1717', fill_type="solid")
            sheet[i][8].value = d[1]
            a += 1
    # except Exception:
    #     print(Exception)
    #     return 1
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['H'].width = 18
    wb.save('/home/Saynir/bot/Ваши_товары.xlsx')


def clear(wb, sheet):
    no_fill = PatternFill(fill_type=None)
    side = Side(border_style=None)
    no_border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    for i in range(2, sheet.max_row+1):
        for x in range(9):
            sheet[i][x].value = None
            sheet[i][x].fill = no_fill
            sheet[i][x].border = no_border
    wb.save('/home/Saynir/bot/Ваши_товары.xlsx')


def prepare_data(seq):
    from functools import reduce
    str_to_send = ''
    for i in seq:
        str_to_send += f'Номер: {i[1]}\n'
        str_to_send += f'Дата: {i[2]}\n'
        str_to_send += f'Код: {i[3]}\n'
        str_to_send += f'Трек: {i[4]}\n'
        str_to_send += f'Вес: {i[5]}\n'
        str_to_send += f'Цена: {i[6]}\n'
        str_to_send += f'Расход: {i[7]}\n'
        str_to_send += f'Сумма: {i[8]}\n'
        str_to_send += f'Статус: {i[9]}\n'
        str_to_send += '----------------------------\n'
    if len(str_to_send) > 4095:
        products = str_to_send.split('----------------------------')
        lst_to_send = []
        for i in range(1, len(products)+1, 20):
            lst_to_send.append(reduce(lambda a, b: a+b, products[:i]))
        return lst_to_send
    if len(str_to_send) == 0:
        str_to_send = 'Товаров с таким кодом нет!'
    return str_to_send


